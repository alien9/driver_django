import os
import zipfile
import tempfile
import time
import re
import pytz
from celery.utils.log import get_logger
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

from django.conf import settings
from django.contrib.auth.models import User

from celery import shared_task
from celery.utils.log import get_task_logger

from django_redis import get_redis_connection

from data.models import DriverRecord

from driver_auth.permissions import is_admin_or_writer
from constance import config

logger = get_task_logger(__name__)
local_tz = pytz.timezone(settings.TIME_ZONE)

logger = get_logger(__name__)

def _utf8(value):
    """
    Helper for properly encoding values that may contain unicode characters.
    From https://github.com/azavea/django-queryset-csv/blob/master/djqscsv/djqscsv.py#L174

    :param value: The string to encode
    """

    if isinstance(value, str):
        return value
    else:
        return str(value)


def _sanitize(value):
    """
    Helper for sanitizing the record type label to ensure it doesn't contain characters that are
    invalid in filenames such as slashes.
    This keeps spaces, periods, underscores, and all unicode characters.

    :param value: The string to sanitize
    """
    return ''.join(char for char in value if char.isalnum() or char in [' ', '.', '_']).rstrip()


@shared_task(track_started=True)
def export_xlsx(query_key, user_id):
    """Exports a set of records to multiple XLSX worksheets in a single file
    :param query_key: A UUID corresponding to a cached SQL query which will be used to filter
                      which records are returned. This is the same key used to generate filtered
                      Windshaft tiles so that the XLSX will correspond to the filters applied in
                      the UI.
    """
    logger.info("Starting export_xlsx")
    # Get Records
    records = get_queryset_by_key(query_key)
    # Get the most recent Schema for the Records' RecordType
    # This assumes that all of the Records have the same RecordType.
    try:
        record_type = records[0].schema.record_type
        schema = record_type.get_current_schema()
    except IndexError:
        raise Exception('Filter includes no records')

    # Get user
    user = User.objects.get(pk=user_id)
    # Create XLSX writer from Schema
    record_writer = DriverRecordExporter(schema)

    # Write records to workbook
    for rec in records:
        record_writer.write_record(rec)
    record_writer.finish()

    # Create single XLSX file with multiple sheets
    filename = "{}-{}.xlsx".format(_utf8(_sanitize(record_type.plural_label)), query_key[:8])
    path = os.path.join(settings.CELERY_EXPORTS_FILE_PATH, filename)

    # Save the consolidated workbook
    record_writer.save_consolidated_workbook(path)

    # Cleanup
    record_writer.cleanup()
    aggregate_xlsx(path, schema)

    return os.path.basename(path)

def get_related_fields(schema):
    """Returns a list of related fields from the schema"""
    result=[]
    for k in schema.schema['definitions'].keys():
        for kk in schema.schema['definitions'][k]['properties'].keys():
            if 'fieldType' in schema.schema['definitions'][k]['properties'][kk] and schema.schema['definitions'][k]['properties'][kk]['fieldType'] == "reference":
                result.append((k, kk, schema.schema['definitions'][k]['properties'][kk]))
    return result

def get_sql_string_by_key(key):
    """Returns a SQL string from Redis using key
    :param key: A UUID pointing to the SQL string
    """
    # Since the records list endpoint bypasses the Django caching framework, do that here too
    redis_conn = get_redis_connection('default')
    return redis_conn.get(key)


def get_queryset_by_key(key):
    """Returns a queryset by filtering Records using the SQL stored in Redis at key
    :param key: A UUID specifying the SQL string to use
    """
    sql_str = get_sql_string_by_key(key)
    return DriverRecord.objects.raw(sql_str)


class DriverRecordExporter(object):
    """Exports Records matching a schema to multiple XLSX worksheets in a single file"""
    def __init__(self, schema_obj):
        # Detect related info types and set up XLSX Writers as necessary
        self.schema = schema_obj.schema

        # Create a single consolidated workbook
        self.consolidated_workbook = openpyxl.Workbook()
        self.consolidated_workbook.remove(self.consolidated_workbook.active)  # Remove default sheet

        # Make output writers for each sheet
        self.rec_writer = self.make_record_and_details_writer()
        # All non-details related info types
        self.writers = {related: self.make_related_info_writer(related, subschema)
                        for related, subschema in self.schema['definitions'].items()
                        if not subschema.get('details')}

        self.rec_wb_dict, self.wb_dicts = self.setup_worksheets()
        self.write_headers()

    def setup_worksheets(self):
        """Create worksheets in the consolidated workbook"""
        # Create main records sheet
        rec_ws = self.consolidated_workbook.create_sheet('records')
        rec_wb_dict = {'worksheet': rec_ws, 'row': 1}
        
        # Create sheets for related info types
        wb_dicts = {}
        for related in self.writers.keys():
            ws = self.consolidated_workbook.create_sheet(related)
            wb_dicts[related] = {'worksheet': ws, 'row': 1}
        
        return (rec_wb_dict, wb_dicts)

    def write_headers(self):
        """Write XLSX headers to all worksheets"""
        # Write headers to all sheets
        self.rec_writer.write_header(self.rec_wb_dict)
        for related_name, writer in self.writers.items():
            writer.write_header(self.wb_dicts[related_name])

    def finish(self):
        """Finalize all worksheets (no file closing needed for in-memory workbook)"""
        pass

    def cleanup(self):
        """No cleanup needed for in-memory workbook"""
        pass

    def save_consolidated_workbook(self, path):
        """Save the consolidated workbook to file"""
        self.consolidated_workbook.save(path)

    def write_record(self, rec):
        """Pass rec's fields through all writers to output all info to XLSX sheets"""
        # First the constants writer
        self.rec_writer.write_record(rec, self.rec_wb_dict)
        # Next, use the related info writers to output to the appropriate sheets
        for related_name, writer in self.writers.items():
            if related_name in rec.data:
                if writer.is_multiple:
                    for item in rec.data[related_name]:
                        writer.write_related(rec.pk, item, self.wb_dicts[related_name])
                else:
                    writer.write_related(rec.pk, rec.data[related_name],
                                         self.wb_dicts[related_name])

    def make_constants_xlsx_writer(self):
        """Generate a Record Writer capable of writing out the non-json fields of a Record"""
        def render_date(d):
            return d.astimezone(local_tz).strftime('%Y-%m-%d %H:%M:%S')
        
        xlsx_columns = ['record_id', 'timezone', 'created', 'modified', 'occurred_from',
                        'occurred_to', 'lat', 'lon', 'location_text',
                        'city', 'city_district', 'county', 'neighborhood', 'road',
                        'state', 'weather', 'light']
        if config.SHOW_RECORD_CREATOR:
            xlsx_columns.append("created_by")
        
        # Model field from which to get data for each column
        source_fields = {
            'record_id': 'uuid',
            'timezone': None,
            'lat': 'geom',
            'lon': 'geom'
        }

        # Some model fields need to be transformed before they can go into XLSX
        value_transforms = {
            'record_id': lambda uuid: str(uuid),
            'timezone': lambda _: settings.TIME_ZONE,
            'created': render_date,
            'modified': render_date,
            'occurred_from': render_date,
            'occurred_to': render_date,
            'lat': lambda geom: geom.y,
            'lon': lambda geom: geom.x,
        }
        return RecordModelWriter(xlsx_columns, source_fields, value_transforms)

    def make_related_info_writer(self, info_name, info_definition, include_record_id=True):
        """Generate a RelatedInfoExporter capable of writing out a particular related info field
        :param info_definition: The definitions entry providing the sub-schema to write out.
        """
        # Need to drop Media fields; we can't export them to XLSX usefully.
        drop_keys = dict()
        for prop, attributes in info_definition['properties'].items():
            if 'media' in attributes:
                drop_keys[prop] = None
        return RelatedInfoWriter(info_name, info_definition, field_transform=drop_keys,
                                 include_record_id=include_record_id)

    def make_record_and_details_writer(self):
        """Generate a writer to put record fields and details in one XLSX sheet"""
        model_writer = self.make_constants_xlsx_writer()
        details = {key: subschema for key, subschema in self.schema['definitions'].items()
                   if subschema.get('details') is True}
        details_key = list(details.keys())[0]
        details_writer = self.make_related_info_writer(details_key, details[details_key],
                                                       include_record_id=False)
        return ModelAndDetailsWriter(model_writer, details_writer, details_key)


class BaseRecordWriter(object):
    """Base class for some common functions that exporters need"""

    def write_header(self, wb_dict):
        """Write the XLSX header to worksheet"""
        ws = wb_dict['worksheet']
        for col_idx, col_name in enumerate(self.xlsx_columns, 1):
            ws.cell(row=1, column=col_idx, value=_utf8(col_name))
        wb_dict['row'] = 2


class ModelAndDetailsWriter(BaseRecordWriter):
    """Exports records' model fields, and the *Details field, to a single XLSX sheet"""
    def __init__(self, model_writer, details_writer, details_key):
        """Creates a combined writer
        :param model_writer: A RecordModelWriter instance that will be used to write model fields
        :param details_writer: A RelatedInfoWriter instance that will be used to write Details
        """
        self.model_writer = model_writer
        self.details_writer = details_writer
        self.details_key = details_key
        self.xlsx_columns = model_writer.xlsx_columns + details_writer.xlsx_columns

    def write_header(self, wb_dict):
        """Write writer headers to XLSX"""
        ws = wb_dict['worksheet']
        for col_idx, col_name in enumerate(self.xlsx_columns, 1):
            ws.cell(row=1, column=col_idx, value=_utf8(col_name))
        wb_dict['row'] = 2

    def write_record(self, record, wb_dict):
        """Pull data from a record, send to appropriate writers, and write to XLSX"""
        ws = wb_dict['worksheet']
        row = wb_dict['row']
        
        # Get model data
        model_data = self.model_writer.get_record_data(record)
        
        # Get details data if available
        details_data = {}
        if self.details_key in record.data:
            details_data = self.details_writer.get_related_data(record.data[self.details_key])
        
        # Combine and write
        combined_data = {**model_data, **details_data}
        col_idx = 1
        for col_name in self.xlsx_columns:
            value = combined_data.get(col_name, '')
            ws.cell(row=row, column=col_idx, value=_utf8(value))
            col_idx += 1
        
        wb_dict['row'] = row + 1


class RecordModelWriter(BaseRecordWriter):
    """Exports records' model fields to XLSX"""
    def __init__(self, xlsx_columns, source_fields=dict(), value_transforms=dict()):
        """Creates a record exporter
        :param xlsx_columns: List of columns names to write out to the XLSX.
                            E.g. ['latitude', 'longitude']
        :param source_fields: Dictionary mapping column names to the name of the model field where
                              the appropriate value can be found.
                              E.g. {'latitude': 'geom', 'longitude': 'geom'}
                              Pulls from attributes with the same name as the column name by default
        :param value_transforms: Dictionary mapping column names to functions by which to transform
                                 model field values before writing to the XLSX.
                                 E.g. {'latitude': lambda geom: geom.y}
                                 If a field is not included here, it will be used directly
        """
        logger.debug("XLSX Columns: " + str(xlsx_columns))
        self.xlsx_columns = xlsx_columns
        self.source_fields = source_fields
        self.value_transforms = value_transforms

    def write_record(self, record, wb_dict):
        """Pull field data from record object, transform, write to XLSX"""
        ws = wb_dict['worksheet']
        row = wb_dict['row']
        
        col_idx = 1
        for column in self.xlsx_columns:
            model_value = self.get_model_value_for_column(record, column)
            xlsx_val = self.transform_model_value(model_value, column)
            ws.cell(row=row, column=col_idx, value=_utf8(xlsx_val))
            col_idx += 1
        
        wb_dict['row'] = row + 1

    def get_record_data(self, record):
        """Gets all field values for a record"""
        output_data = dict()
        for column in self.xlsx_columns:
            model_value = self.get_model_value_for_column(record, column)
            output_data[column] = self.transform_model_value(model_value, column)
        return output_data

    def get_model_value_for_column(self, record, column):
        """Gets the value from the appropriate model field to populate column"""
        model_field = self.source_fields.get(column, column)
        if model_field is None:
            return None
        return getattr(record, model_field)

    def transform_model_value(self, value, column):
        """Transforms value into an appropriate value for column"""
        val_transform = self.value_transforms.get(column, lambda v: v)
        return val_transform(value)


class RelatedInfoWriter(BaseRecordWriter):
    """Exports related info properties to XLSX"""
    def __init__(self, info_name, info_definition, field_transform=dict(), include_record_id=True):
        # Construct a field name mapping
        self.property_transform = field_transform
        try:
            for prop in info_definition['properties']:
                if prop not in self.property_transform:
                    self.property_transform[prop] = prop
        except KeyError:
            raise ValueError("Related info definition has no 'properties'; can't detect fields")
        self.property_transform['_localId'] = info_name + '_id'
        info_columns = [col for col in list(self.property_transform.values()) if col is not None]
        self.output_record_id = include_record_id
        if self.output_record_id:
            self.xlsx_columns = ['record_id'] + info_columns
        else:
            self.xlsx_columns = info_columns
        self.is_multiple = info_definition.get('multiple', False)

    def write_related(self, record_id, related_info, wb_dict):
        """Transform related_info and write to XLSX"""
        ws = wb_dict['worksheet']
        row = wb_dict['row']
        
        # Transform
        output_data = self.transform_value_keys(related_info.copy())

        # Append record_id
        if self.output_record_id:
            output_data['record_id'] = record_id

        # Write
        col_idx = 1
        for col_name in self.xlsx_columns:
            value = output_data.get(col_name, '')
            ws.cell(row=row, column=col_idx, value=_utf8(value))
            col_idx += 1
        
        wb_dict['row'] = row + 1

    def get_related_data(self, related_info):
        """Get all field values for related info"""
        output_data = self.transform_value_keys(related_info.copy())
        return output_data

    def transform_value_keys(self, related_info):
        """Set incoming values to new keys in output_data based on self.property_transform"""
        output_data = dict()
        for in_key, out_key in self.property_transform.items():
            if out_key is not None:
                try:
                    output_data[out_key] = _utf8(related_info.pop(in_key))
                except KeyError:
                    pass
        return output_data


def aggregate_xlsx(file_path, schema):
    """Open the given .xlsx file, perform left joins across all sheets using 'record_id' as the key,
    create (or replace) a worksheet named 'mahdar_records' with the joined result, and save
    the workbook back to the same file path.

    Args:
        file_path (str or Path): Path to the .xlsx file to process.

    Returns:
        str: The file_path of the saved workbook.
    """
    # Read all sheets into dataframes
    sheets = pd.read_excel(file_path, sheet_name=None)
    if not sheets:
        raise ValueError(f"No sheets found in {file_path}")

    # Determine base sheet (prefer 'records')
    base_name = 'records' if 'records' in sheets else next(iter(sheets.keys()))
    base_df = sheets[base_name].copy()

    # Ensure record_id exists in base
    if 'record_id' not in base_df.columns:
        raise ValueError(f"Base sheet '{base_name}' is missing 'record_id' column")

    agg_df = base_df
    """     for table_name, field_name, field in get_related_fields(schema.schema):
        sheet_1=sheets[table_name]
        sheet_2=sheets[field['watch']['target']]
        if field_name in sheet_1.columns and f"{field_name}_id" in sheet_2.columns:
            sheet_1 = sheet_1.merge(sheet_2, left_on=field_name, right_on=f"{field_name}_id", how='outer', suffixes=("", f"_{field['watch']['target']}"))
            sheets[table_name] = sheet_1
        del sheets[field['watch']['target']]
    """
    # Merge other sheets that contain record_id
    for name, df in sheets.items():
        if name == base_name or name == 'mahdar_records':
            continue
        if 'record_id' in df.columns:
            # Avoid duplicate column names by adding suffix with sheet name
            agg_df = agg_df.merge(df, on='record_id', how='left', suffixes=("", f"_{name}"))

    # Write all original sheets back and add/replace mahdar_records
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
        # If 'mahdar_records' already existed, this will replace it
        agg_df.to_excel(writer, sheet_name='mahdar_records', index=False)

    # Load the workbook and remove all sheets except 'mahdar_records'
    wb = openpyxl.load_workbook(file_path)
    sheets_to_remove = [sheet_name for sheet_name in wb.sheetnames if sheet_name != 'mahdar_records']
    for sheet_name in sheets_to_remove:
        del wb[sheet_name]
    wb.save(file_path)

    return str(file_path)
